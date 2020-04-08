#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@Time    : 2020/4/4 22:25
@Author  : cbz
@Site    : https://github.com/1173710224/brain-computing/blob/cbz
@File    : gg.py
@Software: PyCharm
@Descripe: 
"""

import openpyxl
final_states = [15,18,24,25] + [i for i in range(29,43)] + [i for i in range(49,56)]

class core:
    def __init__(self, dfa, input):
        self.dfa = dfa
        self.input = input

    def parse(self):
        """
        use dfa rules to recognize input.
        return tokens.
        """
        keywords = ['auto', 'struct', 'if', 'else', 'for', 'do', 'while', 'const',
            'int', 'double', 'float', 'long', 'char', 'short', 'unsigned',
            'switch', 'break', 'defalut', 'continue', 'return', 'void', 'static',
            'auto', 'enum', 'register', 'typeof', 'volatile', 'union', 'extern']
        parsion = []
        analyses = []
        tmpword = ''
        tmpstate = 0
        linenum = 0
        for line in self.input:
            linenum += 1
            i = 0
            while i < len(line):
                tmpchar = line[i]
                nextstate = self.move(tmpstate,tmpchar)
                if nextstate == None:
                    if tmpword != '':
                        if mapping.__contains__(tmpstate):
                            token = mapping[tmpstate]
                            if token[len(token) - 2] == "*":
                                token = token.replace('*',tmpword)
                            #turn IDN to keyword
                            if token[1:4] == "IDN" and tmpword in keywords:
                                token = "<"+tmpword+",_>"
                            analyses.append((linenum,tmpword,token))
                            parsion.append(tmpword)
                        else:
                            analyses.append((linenum,tmpword,'error'))
                    tmpword = ''
                    tmpstate = 0
                    if tmpchar != '\n' and tmpchar != ' ':
                        i -= 1
                else:
                    tmpstate = nextstate
                    tmpword += tmpchar
                i += 1
        return analyses,parsion

    def move(self, state, inputChar):
        """
        use dfa and input char to give back next state.
        """

        dfa = self.dfa
        for i in range(2, dfa.max_row + 1):
            # print(i,dfa.cell(row=i, column = 1).value)
            if state == dfa.cell(row=i, column=1).value:
                for j in range(2, dfa.max_column + 1):
                    # print(dfa.cell(row=1,column=j).value)
                    if inputChar in str(dfa.cell(row=1, column=j).value):
                        # print(j)
                        return dfa.cell(row=i, column=j).value
                return None
        print("something is wrong",state,inputChar)


def readExcel(path):
    """
    process excel build a data structure
    """
    wb = openpyxl.load_workbook(path)
    sheet = wb['Sheet1']
    return sheet

# 获取测试程序
# file = open('test.c','r')
# data = file.readlines()
# file.close()

# 获取mapping
mapping = {}
file = open('map.txt','r')
out = file.readlines()
file.close()
for tmp in out:
    tmp = tmp.replace('\n','')
    tmp = tmp.split(' ')
    mapping[int(tmp[0])] = tmp[1]

# a = core(readExcel("dfa.xlsx"), data)
# analyses,ret = a.parse()
# print(analyses)
# print(ret)

def back_parse(inputs):
    data = inputs.split('\n')
    for i in range(len(data)):
        data[i] = data[i] + '\n'
    dfa = readExcel("dfa.xlsx")
    a = core(dfa, data)
    analyses, ret = a.parse()
    return analyses,dfa
